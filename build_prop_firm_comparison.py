from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path("prop_firm_50k_comparison.xlsx")
AS_OF = date(2026, 6, 6)

COLORS = {
    "navy": "17365D",
    "blue": "2F75B5",
    "light_blue": "D9EAF7",
    "green": "70AD47",
    "light_green": "E2F0D9",
    "amber": "FFC000",
    "light_amber": "FFF2CC",
    "red": "C00000",
    "light_red": "F4CCCC",
    "gray": "E7E6E6",
    "dark_gray": "666666",
    "white": "FFFFFF",
}

SOURCES = {
    "Apex rules": "https://support.apextraderfunding.com/hc/en-us/articles/46724640813083-EOD-Evaluations",
    "Apex payouts": "https://support.apextraderfunding.com/hc/en-us/articles/47205823183003-EOD-Payouts",
    "Apex pricing reference": "https://propfirmmatch.com/futures/prop-firm-challenges/apex-trader-funding-standard-eod-1-step-50k",
    "Alpha overview": "https://help.alpha-futures.com/en/articles/9491980-alpha-futures-evaluation-qualified-trader-overview",
    "Alpha pricing": "https://alpha-futures.com/how-it-works",
    "Alpha terms": "https://alpha-futures.com/terms-and-conditions",
    "Tradeify pricing": "https://help.tradeify.co/en/articles/14369021-tradeify-pricing-reference",
    "Tradeify eval": "https://help.tradeify.co/en/articles/12853921-select-evaluation-accounts",
    "Tradeify payouts": "https://help.tradeify.co/en/articles/12853966-select-flex-and-select-daily-payout-policies",
    "Lucid pricing": "https://www.lucidtrading.com/",
    "Lucid product": "https://lucidtrading.com/product/lucidtest-50k/",
    "Lucid payouts": "https://support.lucidtrading.com/en/articles/12890092-lucidpro-payouts",
    "Topstep pricing": "https://help.topstep.com/en/articles/9208217-topstep-pricing",
    "Topstep parameters": "https://help.topstep.com/en/articles/8284197-trading-combine-parameters",
    "Topstep payouts": "https://help.topstep.com/en/articles/8284233-topstep-payout-policy/",
    "TPT pricing reference": "https://funded.now/propfirm/takeprofit-trader/50000",
    "TPT payouts": "https://takeprofittraderhelp.zendesk.com/hc/en-us/articles/15172219527581-PRO-Account-Profit-Split-Withdrawal-Rules",
    "MFFU rules": "https://myfundedfutures.com/plans/flex",
    "MFFU pricing/accounts": "https://help.myfundedfutures.com/en/articles/11802636-traders-evaluation-simplified",
    "Bulenox pricing": "https://bulenox.com/member/signup",
    "Bulenox rules": "https://bulenox.com/help/qualification-account/",
    "Bulenox payouts": "https://bulenox.com/help/master-account/",
    "Earn2Trade pricing": "https://www.earn2trade.com/purchase",
    "Earn2Trade rules": "https://help.earn2trade.com/en/articles/5941958-what-are-the-trader-career-path-rules",
    "Earn2Trade funded": "https://help.earn2trade.com/en/articles/6877574-what-offer-will-i-receive-upon-completing-the-trader-career-path",
    "Elite pricing": "https://shop.elitetraderfunding.com/products/50k-end-of-day-drawdown-evaluation-gift-card",
    "Elite payouts": "https://help.elitetraderfunding.com/help/etf-payout-amounts-by-account-type",
    "Trustpilot Apex": "https://www.trustpilot.com/review/apextraderfunding.com",
    "Trustpilot Alpha": "https://www.trustpilot.com/review/alpha-futures.com",
    "Trustpilot Tradeify": "https://www.trustpilot.com/review/tradeify.co",
    "Trustpilot Lucid": "https://www.trustpilot.com/review/lucidtrading.com",
    "Trustpilot Topstep": "https://www.trustpilot.com/review/topstep.com",
    "Trustpilot TPT": "https://www.trustpilot.com/review/takeprofittrader.com",
    "Trustpilot MFFU": "https://www.trustpilot.com/review/myfundedfutures.com",
    "Trustpilot Earn2Trade": "https://www.trustpilot.com/review/www.earn2trade.com",
}

firms = [
    {
        "company": "Apex Trader Funding",
        "plan": "50K EOD Evaluation",
        "account_size": 50000, "eval_price": 450, "promo_price": 45, "billing": "One-time / 30-day assessment",
        "activation": 119, "funded_monthly": 0, "max_accounts": 20, "five_feasible": "Yes",
        "copy": "Yes, own accounts; avoid prohibited coordination/hedging",
        "platforms": "Rithmic/Tradovate ecosystem; NinjaTrader, TradingView, Quantower, Sierra Chart",
        "markets": "CME/CBOT/NYMEX/COMEX futures; verify product list",
        "target": 3000, "drawdown": 2000, "dll": "$1,000 fixed session DLL", "dd_type": "EOD trailing; enforced next session",
        "min_days": 1, "consistency": "None to pass; 50% at payout", "news": "Allowed, subject to conduct rules",
        "overnight": "Intraday only; flat by close", "weekend": "No", "start_contracts": "6 (official support page table should be rechecked at checkout)",
        "max_contracts": "6", "automation": "Restricted; verify current prohibited-conduct policy",
        "first_payout": "After 5 qualifying days at $250+ each", "payout_days": 5, "min_payout": 500,
        "max_first_payout": 1500, "max_recurring": "Payouts 1-6: $1.5k, $1.5k, $2k, $2.5k, $2.5k, $3k",
        "split": 1.00, "frequency": "Up to weekly", "buffer": "$2,100 safety net; $52,600 balance to request $500",
        "payout_effect": "Only profit above safety net withdrawable; PA closes after payout 6",
        "tax": "Needs verification", "realistic_first": 1000,
        "rating": 4.3, "reviews": 19517, "risk": "Medium",
        "complaints": "Historical payout-review/rule-change complaints; promo/base pricing changes; strict conduct rules",
        "praise": "Large community, broad platform support, 100% current payout split, up to 20 accounts",
        "longevity": "Operating since 2021", "clarity_note": "2026 program redesign; price is volatile and promo-led",
        "scores": (8, 7, 7, 10, 7),
        "sources": ["Apex rules", "Apex payouts", "Apex pricing reference", "Trustpilot Apex"],
        "verify": "Normal $450 and $119 activation are volatile; confirm checkout before purchase.",
    },
    {
        "company": "Alpha Futures", "plan": "50K Standard Evaluation", "account_size": 50000,
        "eval_price": 79, "promo_price": None, "billing": "Monthly", "activation": 149, "funded_monthly": 0,
        "max_accounts": 3, "five_feasible": "No (3 qualified max)", "copy": "Own-account copying may be supported; copying another trader prohibited",
        "platforms": "Tradovate/Rithmic options; platform details need verification", "markets": "CME futures; verify exact list",
        "target": 3000, "drawdown": 2000, "dll": "None eval; 2% daily loss guard qualified", "dd_type": "EOD trailing, locks near starting balance",
        "min_days": 2, "consistency": "50% eval; 40% Standard qualified", "news": "Eval yes; funded varies by plan",
        "overnight": "No; flat by 4:20 PM ET", "weekend": "No", "start_contracts": "Scaling plan",
        "max_contracts": "5 minis / 50 micros", "automation": "Algorithmic support reported; verify copier/bot policy",
        "first_payout": "After 5 winning days of $200+", "payout_days": 5, "min_payout": 200,
        "max_first_payout": 1500, "max_recurring": "$1,500/request on current 50K plan shown",
        "split": 0.90, "frequency": "Up to 4x/month", "buffer": "May withdraw up to 50% of profit/request",
        "payout_effect": "Drawdown protection remains important", "tax": "Needs verification", "realistic_first": 1000,
        "rating": 4.9, "reviews": 3952, "risk": "Medium",
        "complaints": "New proprietary platform issues; plan naming/pricing changes; consistency complexity",
        "praise": "Very high recent support ratings, EOD drawdown, fast support",
        "longevity": "Newer firm; shorter operating history than Topstep/Earn2Trade", "clarity_note": "Current site and help-center plan labels conflict in places",
        "scores": (9, 7, 7, 4, 8), "sources": ["Alpha overview", "Alpha pricing", "Alpha terms", "Trustpilot Alpha"],
        "verify": "Current site displays multiple plan variants. Confirm Standard price, activation fee, and five-account objective.",
    },
    {
        "company": "Tradeify", "plan": "50K Select Evaluation", "account_size": 50000,
        "eval_price": 165, "promo_price": 156.75, "billing": "One-time; 5-account bundle gets 5% off",
        "activation": 0, "funded_monthly": 0, "max_accounts": 5, "five_feasible": "Yes",
        "copy": "Yes across own accounts; no hedging/circumvention", "platforms": "Tradovate, Rithmic, WealthCharts; NinjaTrader, TradingView, Quantower, Sierra Chart",
        "markets": "CME-group futures; verify exact broker product list", "target": 3000, "drawdown": 2000,
        "dll": "None in Select eval/Flex funded", "dd_type": "EOD trailing; locks at $50,100",
        "min_days": 3, "consistency": "40% eval; none on Select Flex funded", "news": "Allowed with restrictions around prohibited conduct",
        "overnight": "Needs verification", "weekend": "No", "start_contracts": "2 minis / 20 micros funded",
        "max_contracts": "4 minis / 40 micros", "automation": "Copiers supported; automation policy needs verification",
        "first_payout": "Select Flex after 5 winning days; Daily path can qualify daily", "payout_days": 5,
        "min_payout": 250, "max_first_payout": 2000, "max_recurring": "Flex: generally 50% of profit; Daily capped $1,000 on 50K",
        "split": 0.90, "frequency": "Daily eligibility depending on selected path", "buffer": "Path-specific; Select Flex lock trigger $52,100",
        "payout_effect": "MLL locks at $50,100 after threshold/payout mechanics", "tax": "Needs verification",
        "realistic_first": 1250, "rating": 4.6, "reviews": 2961, "risk": "Low",
        "complaints": "Some KYC/account closure and support-resolution complaints; rules have evolved quickly",
        "praise": "Fast payouts/support, one-time pricing, permanent five-pack discount, flexible payout path",
        "longevity": "Launched 2024", "clarity_note": "Excellent current pricing/rules documentation",
        "scores": (9, 9, 9, 10, 9), "sources": ["Tradeify pricing", "Tradeify eval", "Tradeify payouts", "Trustpilot Tradeify"],
        "verify": "Overnight and automation specifics should be confirmed for selected broker/path.",
    },
    {
        "company": "Lucid Trading", "plan": "50K LucidPro Evaluation", "account_size": 50000,
        "eval_price": 185, "promo_price": 129.50, "billing": "One-time", "activation": 0, "funded_monthly": 0,
        "max_accounts": 5, "five_feasible": "Yes", "copy": "Own-account copying appears supported; confirm current automation policy",
        "platforms": "NinjaTrader, Tradovate, Rithmic, MotiveWave, Quantower", "markets": "CME-group futures; verify product list",
        "target": 3000, "drawdown": 2000, "dll": "$1,200", "dd_type": "EOD trailing",
        "min_days": 1, "consistency": "None eval; 40% funded payout cycle", "news": "Needs verification by plan",
        "overnight": "Needs verification", "weekend": "No", "start_contracts": "4 minis / 40 micros",
        "max_contracts": "4 minis / 40 micros", "automation": "Needs verification in writing",
        "first_payout": "Any day after $500 cycle profit, 40% consistency, and buffer", "payout_days": 3,
        "min_payout": 500, "max_first_payout": 2000, "max_recurring": "$2,500 from payout 2 onward",
        "split": 0.90, "frequency": "On demand; no fixed window", "buffer": "$52,100; payout only above buffer",
        "payout_effect": "Payout reduces cushion above locked loss floor", "tax": "Needs verification",
        "realistic_first": 1250, "rating": 4.7, "reviews": 4125, "risk": "Medium",
        "complaints": "Some closures/automation disputes and support delays; company is relatively new",
        "praise": "Very fast payout reports, clean dashboard, clear payout article, no activation fee",
        "longevity": "Operating since 2025", "clarity_note": "Good help center; checkout pricing/promos change",
        "scores": (8, 8, 9, 9, 8), "sources": ["Lucid pricing", "Lucid product", "Lucid payouts", "Trustpilot Lucid"],
        "verify": "Obtain written confirmation for bots/copiers, news and overnight policy.",
    },
    {
        "company": "Topstep", "plan": "50K Standard Trading Combine", "account_size": 50000,
        "eval_price": 49, "promo_price": None, "billing": "Monthly", "activation": 149, "funded_monthly": 0,
        "max_accounts": 5, "five_feasible": "Yes (5 active XFA max; verify purchase limits)",
        "copy": "TopstepX trade copier available; strict no cross-account hedging", "platforms": "TopstepX; supported third-party availability varies",
        "markets": "Major CME-group futures supported", "target": 3000, "drawdown": 2000,
        "dll": "Optional soft DLL discount; no required DLL in Combine", "dd_type": "Trailing Maximum Loss Limit based on balance",
        "min_days": 2, "consistency": "50% Combine; choose Standard or 40% Consistency XFA",
        "news": "Generally allowed; follow permitted trading hours", "overnight": "No; intraday futures hours",
        "weekend": "No", "start_contracts": "Scaling plan in funded", "max_contracts": "5 minis / 50 micros in Combine",
        "automation": "Trade copier supported; automation subject to prohibited-conduct rules",
        "first_payout": "Standard: 5 days $150+; Consistency: 3 days plus 40%", "payout_days": 5,
        "min_payout": 125, "max_first_payout": 2000, "max_recurring": "$2,000 Standard / $3,000 Consistency per request on new 50K",
        "split": 0.90, "frequency": "Weekly; daily after 30 winning live days", "buffer": "Can request 50% of account balance/profit subject to cap",
        "payout_effect": "Payout can reduce buying power and account cushion", "tax": "Contractor payout process; verify current tax provider",
        "realistic_first": 1250, "rating": 3.5, "reviews": 14166, "risk": "Low",
        "complaints": "Strict hedging enforcement, subscription billing complaints, platform incidents",
        "praise": "12+ year history, real live progression, strong education, clear official help center",
        "longevity": "Founded 2012", "clarity_note": "Best-in-class official documentation, though rules update often",
        "scores": (8, 8, 8, 9, 9), "sources": ["Topstep pricing", "Topstep parameters", "Topstep payouts", "Trustpilot Topstep"],
        "verify": "Confirm current XFA count and selected payout path immediately before purchase.",
    },
    {
        "company": "Take Profit Trader", "plan": "50K Test", "account_size": 50000,
        "eval_price": 170, "promo_price": 85, "billing": "Monthly", "activation": 130, "funded_monthly": 0,
        "max_accounts": 5, "five_feasible": "Yes", "copy": "Yes across own accounts; no hedging",
        "platforms": "Rithmic/Tradovate; NinjaTrader, TradingView, Quantower and others", "markets": "CME-group futures",
        "target": 3000, "drawdown": 2000, "dll": "None on current Test; confirm at checkout", "dd_type": "EOD trailing in Test; PRO mechanics tighten around buffer",
        "min_days": 5, "consistency": "50% Test; none PRO", "news": "Allowed", "overnight": "No",
        "weekend": "No", "start_contracts": "Scaling plan", "max_contracts": "6 minis / 60 micros",
        "automation": "Copiers allowed; bots/EAs restricted", "first_payout": "Day one after reaching $52,000 buffer",
        "payout_days": 1, "min_payout": 0, "max_first_payout": 5000,
        "max_recurring": "No stated cycle cap above buffer; firm may move trader to PRO+ live",
        "split": 0.80, "frequency": "Daily/on demand", "buffer": "$2,000 buffer; normal withdrawals only once balance reaches $52,000",
        "payout_effect": "Withdrawing above buffer preserves floor; inside-buffer withdrawal generally terminates account",
        "tax": "Needs verification", "realistic_first": 1500, "rating": 4.4, "reviews": 9562, "risk": "Medium",
        "complaints": "PRO+ transition/support delays; conflicting third-party drawdown figures; monthly cost",
        "praise": "Daily withdrawals, long payout history, responsive support, clear core buffer concept",
        "longevity": "Operating since 2022", "clarity_note": "Official help center clear on payouts; public pricing page less searchable",
        "scores": (6, 8, 10, 9, 8), "sources": ["TPT pricing reference", "TPT payouts", "Trustpilot TPT"],
        "verify": "Confirm current $2,000 drawdown, $130 activation, and PRO+ transition policy.",
    },
    {
        "company": "MyFundedFutures", "plan": "50K Flex", "account_size": 50000,
        "eval_price": 153, "promo_price": 77, "billing": "Monthly", "activation": 0, "funded_monthly": 0,
        "max_accounts": 3, "five_feasible": "No (new 50K Flex sim-funded cap is 3)",
        "copy": "Own-account copier supported subject to anti-hedging rules", "platforms": "NinjaTrader, TradingView, Quantower, Tradovate and partner platforms",
        "markets": "CME/CBOT/NYMEX/COMEX futures", "target": 3000, "drawdown": 2000,
        "dll": "None (optional lower-priced DLL add-on variant)", "dd_type": "EOD trailing in eval and Flex sim-funded",
        "min_days": 2, "consistency": "50% eval only; none payout stage", "news": "Allowed",
        "overnight": "Needs verification", "weekend": "No", "start_contracts": "2 minis / 20 micros",
        "max_contracts": "5 minis / 50 micros after scaling", "automation": "Copiers supported; bot policy needs verification",
        "first_payout": "5 days at $150+ net", "payout_days": 5, "min_payout": 500,
        "max_first_payout": 2000, "max_recurring": "50% of net profit up to $2,000/cycle; max 5 sim payouts",
        "split": 0.80, "frequency": "Weekly/performance based", "buffer": "No buffer requirement",
        "payout_effect": "After first payout max loss becomes fixed at $100", "tax": "Needs verification",
        "realistic_first": 1000, "rating": 4.9, "reviews": 18858, "risk": "Medium",
        "complaints": "Frequent plan changes; strict post-payout $100 loss floor; KYC/support complaints",
        "praise": "No payout buffer, strong support ratings, clear current plan page, path to live after 5 payouts",
        "longevity": "Operating since 2023", "clarity_note": "Current pages are detailed but product lineup changes frequently",
        "scores": (8, 9, 8, 3, 8), "sources": ["MFFU rules", "MFFU pricing/accounts", "Trustpilot MFFU"],
        "verify": "Not suitable for five funded 50K Flex accounts under the current 3-account cap.",
    },
    {
        "company": "Bulenox", "plan": "50K Option 2 EOD/Scaling", "account_size": 50000,
        "eval_price": 175, "promo_price": None, "billing": "Monthly", "activation": 148, "funded_monthly": 0,
        "max_accounts": 3, "five_feasible": "Needs verification; post-pass gating reported",
        "copy": "Needs verification", "platforms": "Rithmic ecosystem; platform list shown by icons, verify exact support",
        "markets": "Broad CME-group futures list", "target": 3000, "drawdown": 2500,
        "dll": "$1,100 soft daily suspension on Option 2", "dd_type": "EOD trailing/scaling",
        "min_days": 5, "consistency": "Consistency applies in Master payout stage; exact current % needs verification",
        "news": "Needs verification", "overnight": "Limited micros only on legacy help text", "weekend": "No",
        "start_contracts": "2 minis", "max_contracts": "7 after scaling", "automation": "Needs verification",
        "first_payout": "After 10 individual trading days", "payout_days": 10, "min_payout": 1000,
        "max_first_payout": 1500, "max_recurring": "$1,500 for first 3 payouts; later rules/live consolidation apply",
        "split": 1.00, "frequency": "Weekly processing on Wednesday", "buffer": "$2,600 safety reserve",
        "payout_effect": "Reserve remains; after 3 payouts accounts may consolidate to one live account",
        "tax": "1099-MISC/W-8BEN stated on help page", "realistic_first": 1000,
        "rating": None, "reviews": None, "risk": "High",
        "complaints": "Outdated help text, vague strategy enforcement reports, post-pass/live consolidation complexity",
        "praise": "100% first $10k stated, $2.5k drawdown, broad instruments",
        "longevity": "Operating for several years; exact founding date needs verification", "clarity_note": "Public documentation appears dated and internally inconsistent",
        "scores": (5, 6, 5, 4, 4), "sources": ["Bulenox pricing", "Bulenox rules", "Bulenox payouts"],
        "verify": "High-priority verification: account cap, consistency %, copy trading, and post-third-payout consolidation.",
    },
    {
        "company": "Earn2Trade", "plan": "TCP50", "account_size": 50000,
        "eval_price": 190, "promo_price": None, "billing": "Monthly", "activation": 139, "funded_monthly": 0,
        "max_accounts": 1, "five_feasible": "No practical five-account stack; verify partner allocation",
        "copy": "No", "platforms": "Rithmic/Tradovate; NinjaTrader, Finamark, TradingView options",
        "markets": "CME, CBOT, NYMEX, COMEX futures", "target": 3000, "drawdown": 2000,
        "dll": "$1,100", "dd_type": "EOD drawdown, but open/closed equity monitored intraday",
        "min_days": 10, "consistency": "Consistency required; exact formula needs verification",
        "news": "Approved trading times only", "overnight": "No", "weekend": "No",
        "start_contracts": "Progression ladder", "max_contracts": "Up to 6", "automation": "Copy traders prohibited; automation needs approval",
        "first_payout": "Funded withdrawals over $100; first must also cover $139 setup", "payout_days": 1,
        "min_payout": 100, "max_first_payout": 3000, "max_recurring": "No simple cycle cap; TCP scaling model",
        "split": 0.80, "frequency": "On request through partner/Rise", "buffer": "First profit must cover $100 minimum + $139 setup",
        "payout_effect": "Withdraw profit target to advance to larger account in TCP", "tax": "Rise/contractor onboarding; verify forms",
        "realistic_first": 1000, "rating": 4.7, "reviews": 4828, "risk": "Low",
        "complaints": "More restrictive rules, 10-day test, no copier, partner/data/setup complexity",
        "praise": "Founded 2016, transparent pass/withdrawal statistics, education, genuine live-account option",
        "longevity": "Founded 2016", "clarity_note": "Very transparent disclosures, but not designed for account stacking",
        "scores": (5, 5, 6, 1, 9), "sources": ["Earn2Trade pricing", "Earn2Trade rules", "Earn2Trade funded", "Trustpilot Earn2Trade"],
        "verify": "Not a fit for five copied accounts. Confirm partner funding/account-allocation policy.",
    },
    {
        "company": "Elite Trader Funding", "plan": "50K EOD Evaluation", "account_size": 50000,
        "eval_price": 295, "promo_price": None, "billing": "Monthly", "activation": 0, "funded_monthly": 80,
        "max_accounts": 5, "five_feasible": "Needs verification", "copy": "Needs verification",
        "platforms": "NinjaTrader/Rithmic ecosystem; free license/data stated", "markets": "CME-group futures; verify list",
        "target": 3000, "drawdown": 2000, "dll": "$1,100", "dd_type": "EOD trailing",
        "min_days": 5, "consistency": "Needs verification", "news": "Needs verification",
        "overnight": "Needs verification", "weekend": "No", "start_contracts": "8 minis / 80 micros",
        "max_contracts": "8 minis / 80 micros; no scaling in evaluation", "automation": "Needs verification",
        "first_payout": "Active Trade Day cycles; exact 50K EOD table needs verification", "payout_days": 10,
        "min_payout": 500, "max_first_payout": 1500, "max_recurring": "$25,000 aggregate trader cap before live transition stated",
        "split": 0.90, "frequency": "By payout cycle", "buffer": "Needs verification",
        "payout_effect": "Needs verification", "tax": "Needs verification", "realistic_first": 750,
        "rating": None, "reviews": None, "risk": "High",
        "complaints": "Complex product catalog and payout-cycle rules; current public tables difficult to extract",
        "praise": "High contract limit, EOD option, established platform/data package",
        "longevity": "Operating several years", "clarity_note": "Current public information is fragmented",
        "scores": (3, 6, 5, 6, 4), "sources": ["Elite pricing", "Elite payouts"],
        "verify": "Verify current checkout price, account cap, payout cycle/caps, split, copier and funded monthly fee.",
    },
]


def header_style(cell):
    cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    cell.font = Font(color=COLORS["white"], bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def title(ws, text, subtitle=None):
    ws.merge_cells("A1:J1")
    ws["A1"] = text
    ws["A1"].font = Font(size=20, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells("A2:J2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color=COLORS["dark_gray"])


def add_table(ws, start_row, end_row, end_col, name):
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def fit(ws, widths=None):
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    else:
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(c.value or "")) for c in column_cells[:80])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 42)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_source_comment_text(firm):
    return "\n".join(SOURCES[s] for s in firm["sources"])


wb = Workbook()
wb.remove(wb.active)

# Full Comparison
ws = wb.create_sheet("Full Comparison")
title(ws, "50K Futures Prop Firm Full Comparison", f"Public information researched as of {AS_OF.isoformat()}. Normal prices drive scoring; promos are separate.")
headers = [
    "Company", "Plan", "Account Size", "Normal Eval Price", "Promo / 5-Pack Price", "Billing",
    "Activation Fee", "Funded Monthly Fee", "Max Accounts", "Five Funded Feasible?", "Copy Trading",
    "Platforms", "Markets", "Profit Target", "Max Drawdown", "Daily Loss Limit", "Drawdown Type",
    "Min Eval Days", "Consistency", "News Trading", "Overnight", "Weekend", "Starting Contracts",
    "Max Contracts", "Automation / Bots", "First Payout Eligibility", "Payout Days", "Min Payout",
    "Max First Payout / Acct", "Recurring Payout Cap", "Profit Split", "Payout Frequency", "Buffer",
    "Payout Effect", "Tax / Contractor", "Trustpilot", "Review Count", "Risk Grade",
    "Common Complaints", "Common Praise", "Longevity", "Transparency Note", "Needs Verification",
    "Primary Sources"
]
row0 = 4
for c, h in enumerate(headers, 1):
    ws.cell(row0, c, h)
    header_style(ws.cell(row0, c))
for r, f in enumerate(firms, row0 + 1):
    values = [
        f["company"], f["plan"], f["account_size"], f["eval_price"], f["promo_price"], f["billing"],
        f["activation"], f["funded_monthly"], f["max_accounts"], f["five_feasible"], f["copy"],
        f["platforms"], f["markets"], f["target"], f["drawdown"], f["dll"], f["dd_type"], f["min_days"],
        f["consistency"], f["news"], f["overnight"], f["weekend"], f["start_contracts"], f["max_contracts"],
        f["automation"], f["first_payout"], f["payout_days"], f["min_payout"], f["max_first_payout"],
        f["max_recurring"], f["split"], f["frequency"], f["buffer"], f["payout_effect"], f["tax"],
        f["rating"], f["reviews"], f["risk"], f["complaints"], f["praise"], f["longevity"],
        f["clarity_note"], f["verify"], add_source_comment_text(f)
    ]
    for c, value in enumerate(values, 1):
        ws.cell(r, c, value)
    ws.cell(r, 3).number_format = '$#,##0'
    for c in (4, 5, 7, 8, 14, 15, 28, 29):
        ws.cell(r, c).number_format = '$#,##0.00'
    ws.cell(r, 31).number_format = '0%'
    ws.cell(r, 36).number_format = '0.0'
    if f["risk"] == "High":
        ws.cell(r, 38).fill = PatternFill("solid", fgColor=COLORS["light_red"])
    elif f["risk"] == "Low":
        ws.cell(r, 38).fill = PatternFill("solid", fgColor=COLORS["light_green"])
add_table(ws, row0, row0 + len(firms), len(headers), "FullComparisonTable")
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{row0}:{get_column_letter(len(headers))}{row0 + len(firms)}"
ws.conditional_formatting.add(
    f"A{row0+1}:{get_column_letter(len(headers))}{row0+len(firms)}",
    FormulaRule(formula=[f'ISNUMBER(SEARCH("Needs verification",A{row0+1}))'], fill=PatternFill("solid", fgColor=COLORS["light_amber"]))
)
fit(ws, {"A": 23, "B": 25, "C": 13, "D": 16, "E": 18, "F": 20, "G": 15, "H": 17,
         "I": 12, "J": 20, "K": 35, "L": 42, "M": 35, "N": 14, "O": 14, "P": 27,
         "Q": 30, "R": 13, "S": 30, "T": 25, "U": 20, "V": 12, "W": 22, "X": 22,
         "Y": 30, "Z": 35, "AA": 13, "AB": 14, "AC": 21, "AD": 34, "AE": 13,
         "AF": 25, "AG": 34, "AH": 32, "AI": 20, "AJ": 12, "AK": 14, "AL": 12,
         "AM": 42, "AN": 42, "AO": 22, "AP": 35, "AQ": 45, "AR": 60})

# Cost Model
ws = wb.create_sheet("Five Account Cost Model")
title(ws, "Five Account Cost Model", "Assumes all five evaluations pass in one billing period with no resets. Trading-profit buffers are not cash fees.")
cost_headers = [
    "Company", "Plan", "Normal Eval x5", "Promo / Bundle Eval x5", "Monthly Recurring x5",
    "Activation x5", "Funded Monthly x5", "Cost to Funded (1 month)", "Est. Cash Cost to First Payout",
    "Trading Profit Needed to Unlock / Acct", "Max First Payout x5 (Gross)", "Max First Payout x5 (Net)",
    "Realistic First Payout x5 (Net)", "Break-Even Cash Amount", "Five Funded Feasible?", "Notes"
]
row0 = 4
for c, h in enumerate(cost_headers, 1):
    ws.cell(row0, c, h); header_style(ws.cell(row0, c))
for idx, f in enumerate(firms, row0 + 1):
    ws.cell(idx, 1, f["company"]); ws.cell(idx, 2, f["plan"])
    ws.cell(idx, 3, f"='Full Comparison'!D{idx}*5")
    if f["promo_price"] is not None:
        ws.cell(idx, 4, f"='Full Comparison'!E{idx}*5")
    ws.cell(idx, 5, f"=IF('Full Comparison'!F{idx}=\"Monthly\",'Full Comparison'!D{idx}*5,0)")
    ws.cell(idx, 6, f"='Full Comparison'!G{idx}*5")
    ws.cell(idx, 7, f"='Full Comparison'!H{idx}*5")
    ws.cell(idx, 8, f"=C{idx}+F{idx}")
    ws.cell(idx, 9, f"=H{idx}+G{idx}")
    # Profit needed: explicit buffer/min balance approximation from current rules.
    profit_needed = {
        "Apex Trader Funding": 2600, "Alpha Futures": 2000, "Tradeify": 2100,
        "Lucid Trading": 2600, "Topstep": 2500, "Take Profit Trader": 2000,
        "MyFundedFutures": 1000, "Bulenox": 3600, "Earn2Trade": 239,
        "Elite Trader Funding": 2500,
    }[f["company"]]
    ws.cell(idx, 10, profit_needed)
    ws.cell(idx, 11, f"='Full Comparison'!AC{idx}*5")
    ws.cell(idx, 12, f"=K{idx}*'Full Comparison'!AE{idx}")
    ws.cell(idx, 13, f["realistic_first"] * 5 * f["split"])
    ws.cell(idx, 14, f"=I{idx}")
    ws.cell(idx, 15, f["five_feasible"])
    ws.cell(idx, 16, f["verify"])
    for c in range(3, 15):
        ws.cell(idx, c).number_format = '$#,##0.00'
add_table(ws, row0, row0 + len(firms), len(cost_headers), "FiveAccountCostTable")
ws.freeze_panes = "A5"
ws.conditional_formatting.add(f"I5:I{4+len(firms)}", ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"))
ws.conditional_formatting.add(f"M5:M{4+len(firms)}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
fit(ws, {"A": 23, "B": 25, "C": 17, "D": 20, "E": 19, "F": 16, "G": 18, "H": 22,
         "I": 25, "J": 26, "K": 25, "L": 24, "M": 27, "N": 20, "O": 22, "P": 55})

# Rules Matrix
ws = wb.create_sheet("Rules Matrix")
title(ws, "Rules Matrix", "Side-by-side evaluation and funded-rule comparison.")
rule_headers = ["Company", "Plan", "Drawdown Type", "Profit Target", "Max Drawdown", "Daily Loss Limit",
                "Min Days", "Consistency", "News", "Overnight", "Contract Limits", "Copy Trading", "Automation"]
row0 = 4
for c, h in enumerate(rule_headers, 1):
    ws.cell(row0, c, h); header_style(ws.cell(row0, c))
for r, f in enumerate(firms, row0 + 1):
    vals = [f["company"], f["plan"], f["dd_type"], f["target"], f["drawdown"], f["dll"], f["min_days"],
            f["consistency"], f["news"], f["overnight"], f"{f['start_contracts']} -> {f['max_contracts']}",
            f["copy"], f["automation"]]
    for c, v in enumerate(vals, 1): ws.cell(r, c, v)
    ws.cell(r, 4).number_format = ws.cell(r, 5).number_format = '$#,##0'
add_table(ws, row0, row0 + len(firms), len(rule_headers), "RulesMatrixTable")
ws.freeze_panes = "A5"
fit(ws)

# Payout Matrix
ws = wb.create_sheet("Payout Matrix")
title(ws, "Payout Matrix", "Net payout values depend on each firm's stated profit split.")
payout_headers = ["Company", "First Payout Timing", "Required Days", "Minimum Payout", "Max First Payout",
                  "Recurring Cap", "Profit Split", "Frequency", "Buffer Rules", "Payout Effect / Caps"]
row0 = 4
for c, h in enumerate(payout_headers, 1):
    ws.cell(row0, c, h); header_style(ws.cell(row0, c))
for r, f in enumerate(firms, row0 + 1):
    vals = [f["company"], f["first_payout"], f["payout_days"], f["min_payout"], f["max_first_payout"],
            f["max_recurring"], f["split"], f["frequency"], f["buffer"], f["payout_effect"]]
    for c, v in enumerate(vals, 1): ws.cell(r, c, v)
    ws.cell(r, 4).number_format = ws.cell(r, 5).number_format = '$#,##0'
    ws.cell(r, 7).number_format = '0%'
add_table(ws, row0, row0 + len(firms), len(payout_headers), "PayoutMatrixTable")
ws.freeze_panes = "A5"
fit(ws)

# Reputation
ws = wb.create_sheet("Reputation - Risk")
title(ws, "Reputation / Risk", "Review ratings are snapshots, not proof of payout reliability. Themes combine public reviews and documentation quality.")
rep_headers = ["Company", "Trustpilot", "Reviews", "Common Complaints", "Common Praise", "Rule-Change Risk",
               "Payout-Denial Risk", "Overall Risk", "Longevity", "Transparency", "Review Source"]
row0 = 4
for c, h in enumerate(rep_headers, 1):
    ws.cell(row0, c, h); header_style(ws.cell(row0, c))
for r, f in enumerate(firms, row0 + 1):
    change_risk = "High" if f["company"] in ("Apex Trader Funding", "MyFundedFutures", "Bulenox", "Elite Trader Funding") else "Medium"
    denial_risk = "High" if f["company"] in ("Bulenox", "Elite Trader Funding") else ("Medium" if f["risk"] == "Medium" else "Low")
    review_key = next((s for s in f["sources"] if s.startswith("Trustpilot")), None)
    vals = [f["company"], f["rating"], f["reviews"], f["complaints"], f["praise"], change_risk,
            denial_risk, f["risk"], f["longevity"], f["clarity_note"], SOURCES.get(review_key, "Needs verification")]
    for c, v in enumerate(vals, 1): ws.cell(r, c, v)
    if review_key:
        ws.cell(r, 11).hyperlink = SOURCES[review_key]
        ws.cell(r, 11).style = "Hyperlink"
add_table(ws, row0, row0 + len(firms), len(rep_headers), "ReputationRiskTable")
ws.freeze_panes = "A5"
for col in ("F", "G", "H"):
    ws.conditional_formatting.add(f"{col}5:{col}{4+len(firms)}", CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor=COLORS["light_red"])))
fit(ws, {"A": 23, "B": 12, "C": 13, "D": 48, "E": 48, "F": 18, "G": 19, "H": 15, "I": 27, "J": 40, "K": 55})

# Scores
ws = wb.create_sheet("Scores & Dashboard")
title(ws, "Weighted Scorecard & Dashboard", "Weights: Cost 20%, Rules/Drawdown 20%, Payout 25%, Multi-account 15%, Reputation/Transparency 20%. Inputs scored 1-10.")
score_headers = ["Company", "Cost (1-10)", "Rules (1-10)", "Payout (1-10)", "Multi-Account (1-10)",
                 "Reputation (1-10)", "Cost Weight", "Rules Weight", "Payout Weight", "Multi Weight",
                 "Reputation Weight", "Weighted Score / 100", "Cost to First Payout x5", "Max First Payout x5 Net",
                 "Payout Flexibility"]
row0 = 5
weights = [0.20, 0.20, 0.25, 0.15, 0.20]
for i, w in enumerate(weights, 7):
    ws.cell(3, i, w); ws.cell(3, i).number_format = '0%'
ws["A3"] = "Visible scoring weights ->"
ws["A3"].font = Font(bold=True)
for c, h in enumerate(score_headers, 1):
    ws.cell(row0, c, h); header_style(ws.cell(row0, c))
for r, f in enumerate(firms, row0 + 1):
    ws.cell(r, 1, f["company"])
    for c, score in enumerate(f["scores"], 2): ws.cell(r, c, score)
    for c, w in enumerate(weights, 7): ws.cell(r, c, w)
    ws.cell(r, 12, f"=SUMPRODUCT(B{r}:F{r},G{r}:K{r})*10")
    cost_row = r - 1
    ws.cell(r, 13, f"='Five Account Cost Model'!I{cost_row}")
    ws.cell(r, 14, f"='Five Account Cost Model'!L{cost_row}")
    ws.cell(r, 15, f["scores"][2] * 10)
    ws.cell(r, 12).number_format = '0.0'
    for c in (13, 14): ws.cell(r, c).number_format = '$#,##0'
add_table(ws, row0, row0 + len(firms), len(score_headers), "ScorecardTable")
ws.freeze_panes = "A6"
ws.conditional_formatting.add(f"B6:F{5+len(firms)}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
ws.conditional_formatting.add(f"L6:L{5+len(firms)}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))

bar = BarChart()
bar.type = "bar"; bar.style = 10; bar.title = "Estimated Cash Cost to First Payout: 5 Accounts"
bar.y_axis.title = "Firm"; bar.x_axis.title = "USD"
bar.add_data(Reference(ws, min_col=13, min_row=5, max_row=5 + len(firms)), titles_from_data=True)
bar.set_categories(Reference(ws, min_col=1, min_row=6, max_row=5 + len(firms)))
bar.height = 8; bar.width = 15
ws.add_chart(bar, "Q5")

bar2 = BarChart()
bar2.type = "col"; bar2.style = 10; bar2.title = "Maximum First Payout Across 5 Accounts (Net)"
bar2.y_axis.title = "USD"
bar2.add_data(Reference(ws, min_col=14, min_row=5, max_row=5 + len(firms)), titles_from_data=True)
bar2.set_categories(Reference(ws, min_col=1, min_row=6, max_row=5 + len(firms)))
bar2.height = 8; bar2.width = 15
ws.add_chart(bar2, "Q21")

scatter = ScatterChart()
scatter.title = "Cost vs Payout Flexibility"
scatter.x_axis.title = "Cash Cost to First Payout x5"; scatter.y_axis.title = "Payout Flexibility Score"
xvalues = Reference(ws, min_col=13, min_row=6, max_row=5 + len(firms))
yvalues = Reference(ws, min_col=15, min_row=6, max_row=5 + len(firms))
series = Series(yvalues, xvalues, title="Firms")
scatter.series.append(series); scatter.height = 8; scatter.width = 15
ws.add_chart(scatter, "Q37")
fit(ws, {"A": 23, "B": 14, "C": 15, "D": 15, "E": 21, "F": 20, "G": 13, "H": 13,
         "I": 14, "J": 13, "K": 18, "L": 22, "M": 24, "N": 25, "O": 20})

# Executive summary
ws = wb.create_sheet("Executive Summary", 0)
title(ws, "Executive Summary: 5 x $50K Futures Accounts", f"Research date: {AS_OF.strftime('%B %d, %Y')}. Verify checkout and rules immediately before purchase.")
ws["A4"] = "Recommendation"
header_style(ws["A4"])
ws.merge_cells("B4:J4")
ws["B4"] = ("Best overall for the stated five-account objective: Tradeify Select 50K. It combines a permanent 5-account bundle discount, "
             "one-time pricing, no activation fee, EOD drawdown, a selectable payout path, and current official documentation that is unusually clear.")
ws["B4"].fill = PatternFill("solid", fgColor=COLORS["light_green"])
ws["B4"].font = Font(bold=True)
summary = [
    ("Top 3", "1) Tradeify Select 50K  2) Topstep 50K Standard  3) LucidPro 50K"),
    ("Best low-cost normal path", "Topstep: $49/month + $149 activation per account; Tradeify is stronger if avoiding monthly rebills."),
    ("Best payout-friendly", "Take Profit Trader for daily access after its buffer; Tradeify/Lucid are stronger blends of cost and payout terms."),
    ("Best rules / transparency", "Topstep for mature documentation; Tradeify for concise current product documentation."),
    ("Best for copying 5 accounts", "Tradeify, then Topstep. Confirm copier and anti-hedging details in writing before scaling."),
    ("Highest-risk / verify", "Bulenox and Elite Trader Funding. Apex pricing is also highly promotion-sensitive."),
    ("Not eligible for the goal", "MFFU 50K Flex currently caps new traders at 3 sim-funded Flex accounts; Alpha also states 3 qualified accounts; Earn2Trade is not a stacking model."),
    ("Core red flag", "The advertised $50K is not true risk capital. Practical risk budget is usually the $2K-$2.5K drawdown, and payout withdrawals often shrink that cushion."),
]
for i, (label, text) in enumerate(summary, 6):
    ws.cell(i, 1, label).font = Font(bold=True, color=COLORS["navy"])
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)
    ws.cell(i, 2, text)
    ws.cell(i, 2).fill = PatternFill("solid", fgColor=COLORS["light_blue"] if i % 2 == 0 else COLORS["white"])

ws["A16"] = "Top Firms by Weighted Score"
ws["A16"].font = Font(size=14, bold=True, color=COLORS["navy"])
exec_headers = ["Rank", "Company", "Score / 100", "Cost to First Payout x5", "Max First Payout x5 Net", "Five-Account Fit"]
for c, h in enumerate(exec_headers, 1):
    ws.cell(17, c, h); header_style(ws.cell(17, c))
ranked = sorted(firms, key=lambda f: sum(s*w for s, w in zip(f["scores"], weights)), reverse=True)
for r, f in enumerate(ranked[:6], 18):
    score_row = 6 + firms.index(f)
    cost_row = 5 + firms.index(f)
    ws.cell(r, 1, r - 17)
    ws.cell(r, 2, f["company"])
    ws.cell(r, 3, f"='Scores & Dashboard'!L{score_row}")
    ws.cell(r, 4, f"='Five Account Cost Model'!I{cost_row}")
    ws.cell(r, 5, f"='Five Account Cost Model'!L{cost_row}")
    ws.cell(r, 6, f["five_feasible"])
    ws.cell(r, 3).number_format = '0.0'
    ws.cell(r, 4).number_format = ws.cell(r, 5).number_format = '$#,##0'
ws["A26"] = "Method / Caveats"
ws["A26"].font = Font(size=14, bold=True, color=COLORS["navy"])
caveats = [
    "Normal list price is used for cost scoring. Promo prices are shown separately because codes and discounts change frequently.",
    "Cost to first payout means cash fees paid, assuming a first-month pass and no resets. Required trading profit/buffer is shown separately.",
    "Maximum payout values are current stated cycle caps or a conservative modeling amount where the firm says payouts are uncapped.",
    "Review ratings are snapshots from Trustpilot on the research date and are not treated as proof. Complaint themes are allegations/themes, not adjudicated facts.",
    "All futures trading and evaluation fees involve substantial loss risk. This workbook is research, not financial, tax, or legal advice.",
]
for i, text in enumerate(caveats, 27):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
    ws.cell(i, 1, f"• {text}")
fit(ws, {"A": 25, "B": 24, "C": 15, "D": 24, "E": 26, "F": 25, "G": 14, "H": 14, "I": 14, "J": 14})

# Sources
ws = wb.create_sheet("Source Links")
title(ws, "Source Links", "Official company pages are prioritized. Third-party pricing references are labeled and should be verified at checkout.")
source_headers = ["ID", "Source", "Type", "URL", "Accessed", "Notes"]
row0 = 4
for c, h in enumerate(source_headers, 1):
    ws.cell(row0, c, h); header_style(ws.cell(row0, c))
for r, (name, url) in enumerate(SOURCES.items(), row0 + 1):
    source_type = "Review platform" if name.startswith("Trustpilot") else ("Third-party pricing reference" if "reference" in name.lower() else "Official")
    note = "Volatile; confirm at checkout" if "pricing reference" in source_type.lower() else ""
    vals = [r - row0, name, source_type, url, AS_OF, note]
    for c, v in enumerate(vals, 1): ws.cell(r, c, v)
    ws.cell(r, 4).hyperlink = url; ws.cell(r, 4).style = "Hyperlink"
    ws.cell(r, 5).number_format = "yyyy-mm-dd"
add_table(ws, row0, row0 + len(SOURCES), len(source_headers), "SourceLinksTable")
ws.freeze_panes = "A5"
fit(ws, {"A": 8, "B": 28, "C": 28, "D": 90, "E": 14, "F": 30})

# Global polish
thin = Side(style="thin", color="D9E1F2")
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.auto_filter.ref
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(OUTPUT)

# Reopen to verify the artifact is structurally valid.
check = load_workbook(OUTPUT, data_only=False)
required = {
    "Executive Summary", "Full Comparison", "Five Account Cost Model", "Rules Matrix",
    "Payout Matrix", "Reputation - Risk", "Source Links", "Scores & Dashboard"
}
missing = required.difference(check.sheetnames)
if missing:
    raise RuntimeError(f"Missing sheets: {sorted(missing)}")
if OUTPUT.stat().st_size < 25_000:
    raise RuntimeError("Workbook is unexpectedly small")
print(f"Created {OUTPUT.resolve()} ({OUTPUT.stat().st_size:,} bytes)")
